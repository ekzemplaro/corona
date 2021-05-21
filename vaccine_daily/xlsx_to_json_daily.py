#!/usr/bin/python
# -*- coding: utf-8 -*-
#
#	xlsx_to_json_daily.py
#
#						May/21/2021
#
# -------------------------------------------------------------------
import sys
import json
from openpyxl import load_workbook

# -------------------------------------------------------------------
def file_write_proc(file_name,str_out):
#
	fp_out = open(file_name,mode='w',encoding='utf-8')
	fp_out.write(str_out)
	fp_out.close()
#
# -------------------------------------------------------------------
sys.stderr.write("*** 開始 ***\n")

file_xlsx = sys.argv[1]
file_json = sys.argv[2]
sys.stderr.write(file_xlsx + "\n")
#
wb = load_workbook(filename = file_xlsx)
ws = wb.active

sys.stderr.write("%d\n" % ws.max_column)
sys.stderr.write("%d\n" % ws.max_row)
#
array_out = []
icount = 0
for row in ws.rows:
	if row[2].value != None:
		unit_aa = {}
		unit_aa["date"] = str(row[0].value)
		unit_aa["number"] = row[2].value
		unit_aa["first"] = row[3].value
		unit_aa["second"] = row[4].value
		array_out.append(unit_aa)
#
out_str = json.dumps(array_out)

file_write_proc(file_json,out_str)

# print(out_str)

sys.stderr.write("*** 終了 ***\n")
# -------------------------------------------------------------------
